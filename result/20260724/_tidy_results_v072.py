"""Tidy the v0.72 LEAP demand export into machine-readable CSVs.

Source: mailbox/20260724/v_0.72 Demand Result.xlsx — 3 sheets (BAS/ATS/RAS),
Final Energy Demand, 10 AMS x 56 years (2005-2060), unit Billion Gigajoules.

Two quirks this cycle, both handled here (see README_v0.72_results.md):

  1. MIXED EXPORT SHAPE. 'Demand BAS' and 'Demand ATS' were exported with FLAT
     full branch paths; 'Demand RAS' with an INDENTED tree (3 spaces/level).
     Both are parsed to the same full-path grain.
  2. 100-CHAR PATH TRUNCATION in the flat sheets. LEAP clips the flat branch
     path at exactly 100 characters, so 12 deep Industry leaves per sheet
     arrive with mangled labels ('...Liquid FF\\Residual Fue'). VALUES are
     unaffected. Repaired against the RAS tree roster, which is immune (each
     cell holds only the leaf name): every RAS path longer than 100 chars is
     re-truncated the same way to build the lookup key, so the 99-char cases
     (trailing space stripped, e.g. '...Liquid FF\\Residual') are caught too.
     The 13 keys are mutually unique — asserted below.

  3. SHEET-SHAPE ASYMMETRY (not repairable — reported, see README). Five
     Industry branches are single leaves in BAS/ATS but split one level
     deeper in RAS. Three of them are genuinely coarser in BAS/ATS:
     Cement Kiln Conventional\\Heat (5 fuels in RAS), BOF\\BF (5), EAF\\DRI (2).
     Those rows carry carrier='mixed' and fuel_resolved=False.

Carrier convention carried forward from result/20260709/_tidy_results.py:
thermal/combustion fuels in PJ, electricity carriers in GWh.
"""
import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "mailbox" / "20260724" / "v_0.72 Demand Result.xlsx"
OUT = Path(__file__).resolve().parent
BS = chr(92)

EJ_TO_PJ = 1000.0              # 1 Billion Gigajoule = 1e9 GJ = 1000 PJ
PJ_TO_GWH = 1_000_000 / 3600   # 277.7778 GWh

THERMAL_FUELS = {
    "Coal", "Coal Bituminous", "Coal bituminous", "Coal Sub bituminous",
    "Coal Anthracite", "Coal Lignite", "Coal Unspecified", "Metalurgical Coke",
    "Brown Coal Briquettes", "Hard Coal Briquettes", "Gasoline",
    "Aviation Gasoline", "Blended Gasoline", "Kerosene", "Jet Kerosene",
    "Kerosene and Candles", "Diesel", "Blended Diesel", "Residual Fuel Oil",
    "Refinery Gas", "LPG", "Bitumen", "Petroleum Coke", "Natural Gas", "LNG",
    "Crude Oil", "Oil", "Naphtha", "Lubricants", "Wood", "Efficient Wood",
    "Bagasse", "Charcoal", "Other Biomass", "Biomass", "Biogas",
    "Municipal Solid Waste", "Ethanol", "Biodiesel", "SAF", "Hydrogen",
    "Solar Heating", "Solar",
}
ELECTRIC_DEVICES = {
    "Electricity", "Conventional Electric", "Induction Electric",
    "Incandescent", "CFL", "Fluorescent", "Halogen", "LED",
    "Heat Pump", "Heat Pump Outside Air",
}
EFFICIENCY_CLASSES = {
    "Existing", "Efficient", "Best Practice", "High", "Medium", "Low",
    "High_eff", "Mid_eff", "Low_eff", "Current_Stock Average",
    "Current_Sales Average", "Current Stock_Average", "Current Sales_Average",
    "Colocation", "Enterprise", "Hyperscale",
}


# Quirk 3: BAS/ATS report these Industry routes one level shallower than RAS.
# 'Scrap' resolves cleanly — RAS gives it a single child, Electricity. The
# other three aggregate several fuels and cannot be attributed from this
# export, so they are marked mixed / unresolved rather than guessed at.
ROUTE_AGGREGATES = {"BF", "DRI", "Heat"}


def classify_carrier(leaf, path_parts):
    """-> (carrier, fuel, confident, fuel_resolved) for a demand leaf."""
    if leaf == "Scrap":
        return "electricity", "Electricity", True, True
    if leaf in ROUTE_AGGREGATES:
        return "mixed", f"Unresolved ({leaf} route)", True, False
    if leaf == "Electricity" or "Electric" in leaf or leaf in ELECTRIC_DEVICES:
        return "electricity", "Electricity", True, True
    if "Electricity" in path_parts:
        return "electricity", "Electricity", True, True
    if leaf in THERMAL_FUELS:
        return "thermal", leaf, True, True
    if leaf in EFFICIENCY_CLASSES:
        return "electricity", "Electricity", True, True
    return "thermal", leaf, False, False


def meta_after(cell, key):
    s = str(cell or "")
    return s.split(key, 1)[1].strip() if key in s else ""


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    scenario = meta_after(rows[1][0], "Scenario:").split(",")[0].strip()
    prefix = meta_after(rows[2][0], "Branch:").strip()
    unit = meta_after(rows[3][0], "Units:").strip()
    cols = []
    for j, c in enumerate(rows[5]):
        m = re.match(r"^(.+?)\s+(\d{4})$", str(c or "").strip())
        if m:
            cols.append((j, m.group(1).strip(), int(m.group(2))))
    body = [r for r in rows[6:] if r[0] is not None and str(r[0]).strip()]
    indented = any(str(r[0]).startswith(" ") for r in body)
    out, stack = [], []
    for i, r in enumerate(body):
        s = str(r[0])
        name = s.strip()
        if name == "Total":
            continue
        if indented:
            depth = (len(s) - len(s.lstrip(" "))) // 3
            stack = stack[:depth] + [name]
            nxt = body[i + 1] if i + 1 < len(body) else None
            nxt_depth = ((len(str(nxt[0])) - len(str(nxt[0]).lstrip(" "))) // 3
                         if nxt is not None else -1)
            if nxt_depth > depth:
                continue                       # container branch, not a leaf
            out.append((BS.join(stack), r))
        else:
            out.append((name, r))
    return scenario, prefix, unit, cols, out, indented


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
sheets = {}
for sheet in wb.sheetnames:
    sheets[sheet] = read_sheet(wb[sheet])
    sc, pre, un, cols, leaves, ind = sheets[sheet]
    assert un == "Billion Gigajoules", f"unexpected unit {un!r} on {sheet}"
    print(f"{sheet:<12} {sc:<30} {'TREE' if ind else 'FLAT':<5} "
          f"{len(leaves):>4} leaves  {len(cols):>4} region-year cols")

# ---- repair 100-char truncated paths in the flat sheets ------------------
# Build the lookup by re-truncating every long RAS path exactly as LEAP does,
# so a key matches whether or not char 100 happened to be a space.
ras_roster = sorted({p for p, _ in sheets["Demand RAS"][4]})
trunc_key = {}
for r in ras_roster:
    if len(r) > 100:
        trunc_key.setdefault(r[:100].rstrip(), []).append(r)
dupes = {k: v for k, v in trunc_key.items() if len(v) > 1}
assert not dupes, f"ambiguous truncation keys: {dupes}"
repairs = {k: v[0] for k, v in trunc_key.items()}

fixed_count = Counter()
for sheet, (_, _, _, _, leaves, indented) in sheets.items():
    if indented:
        continue
    stranded = [p for p, _ in leaves
                if len(p) >= 99 and p not in repairs and p not in ras_roster]
    assert not stranded, f"{sheet}: unrepairable long paths {stranded}"
print(f"\ntruncation keys built from RAS: {len(repairs)} (all unique)")

# ---- melt ---------------------------------------------------------------
FIELDS = ["domain", "scenario", "scenario_code", "sector", "subsector", "layer",
          "fuel", "carrier", "region", "year", "value", "unit", "value_pj",
          "confident_carrier", "fuel_resolved", "path_repaired",
          "branch_leaf", "branch_path"]
CODE = {"Baseline Simulation": "BAS", "AMS Target Scenario": "ATS",
        "Regional Aspiration Scenario": "RAS"}

rows_out = []
carrier_map = defaultdict(lambda: [None, Counter(), None])
totals_check = defaultdict(float)
for sheet, (scenario, prefix, unit, cols, leaves, indented) in sheets.items():
    n = 0
    for path, r in leaves:
        was_trunc = path in repairs
        if was_trunc:
            path = repairs[path]
            fixed_count[sheet] += 1
        parts = path.split(BS)
        leaf, sector = parts[-1], parts[0]
        subsector = parts[1] if len(parts) > 1 else ""
        if "Historical" in parts:
            layer = "Historical"
        elif any(p.startswith("Projection") for p in parts):
            layer = "Projection"
        else:
            layer = "Direct"
        carrier, fuel, confident, resolved = classify_carrier(leaf, parts[:-1])
        carrier_map[leaf][0], carrier_map[leaf][2] = carrier, confident
        if not confident:
            carrier_map[leaf][1][BS.join(parts[1:-1])] += 1
        bp = prefix + BS + path
        for (j, region, year) in cols:
            v = r[j] if j < len(r) else None
            try:
                ej = float(v)
            except (TypeError, ValueError):
                continue
            pj = ej * EJ_TO_PJ
            totals_check[(scenario, year)] += ej
            if pj == 0.0:
                continue                        # skip-zeros (§7.4)
            value, uom = ((pj * PJ_TO_GWH, "GWh") if carrier == "electricity"
                          else (pj, "PJ"))
            rows_out.append({
                "domain": "demand", "scenario": scenario,
                "scenario_code": CODE[scenario], "sector": sector,
                "subsector": subsector, "layer": layer, "fuel": fuel,
                "carrier": carrier, "region": region, "year": year,
                "value": round(value, 6), "unit": uom, "value_pj": round(pj, 6),
                "confident_carrier": confident, "fuel_resolved": resolved,
                "path_repaired": was_trunc,
                "branch_leaf": leaf, "branch_path": bp})
            n += 1
    print(f"  {sheet:<12} {n:>7} non-zero rows  ({fixed_count[sheet]} repaired paths)")


def write(path, fields, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


write(OUT / "aeo9_v0.72_demand_tidy.csv", FIELDS, rows_out)
print(f"\naeo9_v0.72_demand_tidy.csv  ->  {len(rows_out):,} rows")

# ---- rollups ------------------------------------------------------------
def rollup(keys, name, extra=None):
    agg = defaultdict(float)
    for d in rows_out:
        agg[tuple(d[k] for k in keys)] += d["value_pj"]
    out = []
    for k, pj in sorted(agg.items()):
        rec = dict(zip(keys, k))
        rec["value_pj"] = round(pj, 6)
        rec["value_twh"] = round(pj / 3.6, 6)
        out.append(rec)
    write(OUT / name, keys + ["value_pj", "value_twh"], out)
    print(f"{name}  ->  {len(out):,} rows")
    return out


rollup(["scenario_code", "sector", "region", "year"],
       "aeo9_v0.72_demand_by_sector.csv")
rollup(["scenario_code", "sector", "subsector", "region", "year"],
       "aeo9_v0.72_demand_by_subsector.csv")
rollup(["scenario_code", "fuel", "carrier", "region", "year"],
       "aeo9_v0.72_demand_by_fuel.csv")
rollup(["scenario_code", "carrier", "region", "year"],
       "aeo9_v0.72_demand_by_carrier.csv")

dc = [d for d in rows_out if "Data_Center" in d["branch_path"]]
write(OUT / "aeo9_v0.72_datacenter.csv", FIELDS, dc)
print(f"aeo9_v0.72_datacenter.csv  ->  {len(dc):,} rows")

# ---- verification -------------------------------------------------------
print("\n=== VERIFY: leaf sum vs each sheet's own Total row ===")
wb2 = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ok = True
for sheet in wb2.sheetnames:
    ws = wb2[sheet]
    rows = list(ws.iter_rows(values_only=True))
    scenario = meta_after(rows[1][0], "Scenario:").split(",")[0].strip()
    hdr = rows[5]
    trow = next(r for r in rows[6:] if r[0] and str(r[0]).strip() == "Total")
    for y in (2020, 2030, 2045, 2060):
        js = [j for j, c in enumerate(hdr) if str(c or "").strip().endswith(f" {y}")]
        src = sum(float(trow[j]) for j in js if isinstance(trow[j], (int, float)))
        mine = totals_check[(scenario, y)]
        flag = "OK " if abs(src - mine) < 1e-6 else "MISMATCH"
        if flag != "OK ":
            ok = False
        print(f"  {flag} {scenario:<30} {y}  sheet={src:12.6f}  leaves={mine:12.6f} EJ")
print("ALL SHEETS RECONCILE" if ok else "*** RECONCILIATION FAILED ***")

print("\n=== carrier classification ===")
unsure = {k: v for k, v in carrier_map.items() if not v[2]}
print("electricity leaves:", sorted(k for k, v in carrier_map.items() if v[0] == "electricity"))
print("\nthermal leaves:", sorted(k for k, v in carrier_map.items() if v[0] == "thermal" and v[2]))
print(f"\nUNCONFIRMED ({len(unsure)}):")
for k, v in sorted(unsure.items()):
    print(f"   {k!r} under {dict(v[1])}")
if not unsure:
    print("   none — every leaf classified confidently")
sys.exit(0 if ok and not unsure else 1)
