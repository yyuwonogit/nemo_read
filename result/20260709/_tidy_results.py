"""Melt v0.71 LEAP wide result exports into tidy machine-readable CSVs.

Supply (power): Generation (TWh) + Capacity (GW), flat tech list, 3 scenarios.
Demand: leaf-level final energy demand, differentiated by fuel. Per user
directive 2026-07-09: thermal/combustion fuels in PJ (native Million GJ);
anything whose energy carrier is ELECTRICITY (the Electricity fuel, and
electric devices/appliances) in GWh (1 PJ = 277.778 GWh, exact).

Carrier classification is explicit and surfaced for validation. `value_pj`
(native) is kept on every demand row for audit/aggregation.
"""
import csv, re
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl

SRC = Path(__file__).resolve().parents[2] / "mailbox" / "20260709"
OUT = Path(__file__).resolve().parent
BS = chr(92)
PJ_TO_GWH = 1_000_000 / 3600  # 1 PJ = 1e6 GJ / 3600 GJ per GWh = 277.7778 GWh

VAR_MAP = {"Outputs by Feedstock Fuel": "Generation", "Capacity": "Capacity",
           "Final Energy Demand": "Final Energy Demand"}
UNIT_CLEAN = {"Thousand Gigawatt-Hours": "TWh", "Thousand Megawatts": "GW",
              "Million Gigajoules": "PJ"}

# ---- demand carrier classification --------------------------------------
THERMAL_FUELS = {
    "Coal", "Coal Bituminous", "Coal bituminous", "Coal Sub bituminous",
    "Coal Anthracite", "Coal Lignite", "Coal Unspecified", "Metalurgical Coke",
    "Brown Coal Briquettes", "Hard Coal Briquettes", "Gasoline",
    "Aviation Gasoline", "Blended Gasoline", "Kerosene", "Jet Kerosene",
    "Kerosene and Candles", "Diesel", "Blended Diesel", "Residual Fuel Oil",
    "Refinery Gas", "LPG", "Bitumen", "Petroleum Coke", "Natural Gas", "LNG",
    "Crude Oil", "Oil", "Naphtha", "Lubricants", "Wood", "Efficient Wood",
    "Bagasse", "Charcoal", "Other Biomass", "Biomass", "Biogas",
    "Municipal Solid Waste", "Ethanol", "Biodiesel", "SAF", "Hydrogen",
    "Solar Heating", "Solar",
}
ELECTRIC_DEVICES = {
    "Electricity", "Conventional Electric", "Induction Electric",
    "Incandescent", "CFL", "Fluorescent", "Halogen", "LED",
    "Heat Pump", "Heat Pump Outside Air",
}
# Appliance efficiency / stock / vintage descriptors: these are terminal
# device classes under an electric appliance end-use (AC, refrigeration,
# washing machine, TV, data-centre, ...). Fuel-driven end-uses (cooking,
# water heating) name their leaf by the actual fuel instead, so a pure
# descriptor leaf is electricity.
EFFICIENCY_CLASSES = {
    "Existing", "Efficient", "Best Practice", "High", "Medium", "Low",
    "High_eff", "Mid_eff", "Low_eff", "Current_Stock Average",
    "Current_Sales Average", "Current Stock_Average", "Current Sales_Average",
    "Colocation", "Enterprise", "Hyperscale",
}

def classify_carrier(leaf, path_parts):
    """-> (carrier, fuel, confident) for a demand leaf."""
    if leaf == "Electricity" or "Electric" in leaf or leaf in ELECTRIC_DEVICES:
        return "electricity", "Electricity", True
    if "Electricity" in path_parts:            # e.g. Lighting\Electricity\LED
        return "electricity", "Electricity", True
    if leaf in THERMAL_FUELS:
        return "thermal", leaf, True
    if leaf in EFFICIENCY_CLASSES:             # electric-appliance stock class
        return "electricity", "Electricity", True
    return "thermal", leaf, False              # UNKNOWN -> flag, default PJ

def meta_after(cell, key):
    s = str(cell or "")
    return s.split(key, 1)[1].strip() if key in s else ""

def load_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    variable = VAR_MAP.get(str(rows[0][0]).strip(), str(rows[0][0]).strip())
    scenario = meta_after(rows[1][0], "Scenario:").split(",")[0].strip()
    prefix = meta_after(rows[2][0], "Branch:").strip()
    unit = meta_after(rows[3][0], "Units:").strip()
    header = rows[5]
    cols = []
    for j, c in enumerate(header):
        m = re.match(r"^(.+?)\s+(\d{4})$", str(c or "").strip())
        if m:
            cols.append((j, m.group(1).strip(), int(m.group(2))))
    parsed = []
    for r in rows[6:]:
        b = r[0]
        if b is None or str(b).strip() == "":
            continue
        s = str(b)
        depth = (len(s) - len(s.lstrip())) // 3
        parsed.append((depth, s.strip(), r))
    return variable, scenario, prefix, unit, cols, parsed

# ---- SUPPLY (power) ------------------------------------------------------
POWER_FIELDS = ["domain", "variable", "scenario", "region", "year", "value",
                "unit", "branch_leaf", "branch_path"]
power_rows = []
wb = openpyxl.load_workbook(SRC / "v_0.71 Power Result.xlsx", read_only=True, data_only=True)
print("POWER (supply):")
for sheet in wb.sheetnames:
    variable, scenario, prefix, unit, cols, parsed = load_sheet(wb[sheet])
    uc = UNIT_CLEAN.get(unit, unit)
    n = 0
    for depth, name, r in parsed:
        if name == "Total":            # LEAP export sum row, not a technology
            continue
        bp = prefix + BS + name
        for (j, region, year) in cols:
            v = r[j] if j < len(r) else None
            if v in (None, "") or (isinstance(v, str) and not v.strip()):
                continue
            try: val = float(v)
            except (TypeError, ValueError): continue
            power_rows.append({"domain": "power", "variable": variable,
                "scenario": scenario, "region": region, "year": year,
                "value": val, "unit": uc, "branch_leaf": name, "branch_path": bp})
            n += 1
    print(f"  {sheet:<10} {variable:<12} {scenario:<28} {uc:<5} {n:>6} rows")

# ---- DEMAND --------------------------------------------------------------
DEMAND_FIELDS = ["domain", "scenario", "sector", "layer", "fuel", "carrier",
                 "region", "year", "value", "unit", "value_pj",
                 "confident_carrier", "branch_leaf", "branch_path"]
demand_rows = []
carrier_map = defaultdict(lambda: [None, Counter(), None])  # leaf -> [carrier, endusesample, confident]
wb = openpyxl.load_workbook(SRC / "v_0.71 Demand Result.xlsx", read_only=True, data_only=True)
print("\nDEMAND:")
for sheet in wb.sheetnames:
    variable, scenario, prefix, unit, cols, parsed = load_sheet(wb[sheet])
    assert unit == "Million Gigajoules", f"unexpected demand unit {unit}"
    path = []
    n = 0
    for i, (depth, name, r) in enumerate(parsed):
        path = path[:depth] + [name]
        is_leaf = (i + 1 >= len(parsed)) or (parsed[i + 1][0] <= depth)
        if not is_leaf:
            continue
        if name == "Total":           # sector-list 'Total' pseudo-row
            continue
        sector = path[0]
        if "Historical" in path:
            layer = "Historical"
        elif any(p.startswith("Projection") for p in path):
            layer = "Projection"
        else:
            layer = "Direct"
        carrier, fuel, confident = classify_carrier(name, path[:-1])
        carrier_map[name][0] = carrier
        carrier_map[name][2] = confident
        if not confident:
            carrier_map[name][1][BS.join(path[1:-1])] += 1
        bp = prefix + BS + BS.join(path)
        for (j, region, year) in cols:
            v = r[j] if j < len(r) else None
            if v in (None, "") or (isinstance(v, str) and not v.strip()):
                continue
            try: pj = float(v)
            except (TypeError, ValueError): continue
            if carrier == "electricity":
                value, uom = pj * PJ_TO_GWH, "GWh"
            else:
                value, uom = pj, "PJ"
            demand_rows.append({"domain": "demand", "scenario": scenario,
                "sector": sector, "layer": layer, "fuel": fuel,
                "carrier": carrier, "region": region, "year": year,
                "value": value, "unit": uom, "value_pj": pj,
                "confident_carrier": confident, "branch_leaf": name,
                "branch_path": bp})
            n += 1
    print(f"  {sheet:<12} {scenario:<28} {n:>7} leaf rows")

def write(path, fields, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields); w.writeheader(); w.writerows(rows)
write(OUT / "aeo9_v0.71_supply_power_tidy.csv", POWER_FIELDS, power_rows)
write(OUT / "aeo9_v0.71_demand_by_fuel_tidy.csv", DEMAND_FIELDS, demand_rows)
print(f"\npower rows={len(power_rows)}  demand rows={len(demand_rows)}")

# ---- classification report ----------------------------------------------
print("\n=== DEMAND carrier classification (validate) ===")
elec = sorted(k for k, v in carrier_map.items() if v[0] == "electricity")
therm = sorted(k for k, v in carrier_map.items() if v[0] == "thermal" and v[2])
unsure = {k: v for k, v in carrier_map.items() if not v[2]}
print(f"ELECTRICITY -> GWh ({len(elec)}): {elec}")
print(f"\nTHERMAL -> PJ ({len(therm)}): {therm}")
print(f"\nUNCONFIRMED device-class leaves (defaulted to THERMAL/PJ — REVIEW) ({len(unsure)}):")
for k, v in sorted(unsure.items()):
    print(f"   {k:<22} under end-uses: {dict(v[1])}")

# ---- verification: PJ conservation (leaves sum to sector aggregate) ------
print("\n=== VERIFY: leaf PJ sums to the source sector-aggregate row ===")
wb = openpyxl.load_workbook(SRC / "v_0.71 Demand Result.xlsx", read_only=True, data_only=True)
ws = wb["Demand RAS"]
rows = list(ws.iter_rows(values_only=True))
hdr = rows[5]
# column for Indonesia 2050
tcol = next(j for j, c in enumerate(hdr) if str(c or "").strip() == "Indonesia 2050")
# source 'Industry' sector-aggregate cell (depth-0 row)
src_industry = None
for r in rows[6:]:
    if r[0] and str(r[0]).strip() == "Industry" and (len(str(r[0])) - len(str(r[0]).lstrip())) == 0:
        src_industry = float(r[tcol]); break
# our leaves under Industry, Indonesia 2050, summed back to PJ
leaf_sum = sum(d["value_pj"] for d in demand_rows
               if d["scenario"] == "Regional Aspiration Scenario"
               and d["sector"] == "Industry" and d["region"] == "Indonesia"
               and d["year"] == 2050)
print(f"  Industry / Indonesia / 2050 (RAS):  source sector row = {src_industry:.3f} PJ  |  "
      f"our leaf sum = {leaf_sum:.3f} PJ  ->  {'MATCH' if abs(src_industry-leaf_sum)<0.5 else 'DIFF (hierarchy note)'}")
print("  (small diffs are expected if a sector mixes Historical+Projection layers; leaves are the finest grain)")
elec_gwh = sum(d["value"] for d in demand_rows if d["carrier"] == "electricity"
               and d["scenario"] == "Regional Aspiration Scenario" and d["year"] == 2050)
elec_pj = sum(d["value_pj"] for d in demand_rows if d["carrier"] == "electricity"
              and d["scenario"] == "Regional Aspiration Scenario" and d["year"] == 2050)
print(f"  electricity 2050 RAS: {elec_pj:.1f} PJ == {elec_gwh:,.0f} GWh  (x277.778 check: {elec_pj*PJ_TO_GWH:,.0f})")
