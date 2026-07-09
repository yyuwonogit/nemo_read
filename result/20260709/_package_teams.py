"""Build v0.71 delivery packages for all remaining sector teams, to the same
5-part shape as the residential/power ships: results / input / leap_structure /
connected_drivers / full_results. Also tidies the Resources result.
"""
import csv, re, os, zipfile, shutil
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parents[2]
MB = REPO / "mailbox" / "20260709"
R = REPO / "result" / "20260709"
H = REPO / "inject"          # inject/<team>/structure_handover_20260703
OUTBOX = REPO / "outbox"
BS = chr(92)
AMS10 = ["Brunei","Cambodia","Indonesia","Laos","Malaysia","Myanmar",
         "Philippines","Singapore","Thailand","Vietnam"]

# ---------- tidy the Resources result (Primary Supply EJ + Exports PJ) ----------
def meta_after(c, k):
    s = str(c or ""); return s.split(k,1)[1].strip() if k in s else ""
res_rows = []
wb = openpyxl.load_workbook(MB / "v_0.71 Resources Result.xlsx", read_only=True, data_only=True)
for sheet in wb.sheetnames:
    rows = list(wb[sheet].iter_rows(values_only=True))
    variable = str(rows[0][0]).strip()            # 'Exports' or 'Primary Supply'
    scenario = meta_after(rows[1][0], "Scenario:").split(",")[0].strip()
    unit = meta_after(rows[3][0], "Units:").strip()
    hdr = rows[5]
    cols = [(j, m.group(1).strip(), int(m.group(2)))
            for j,c in enumerate(hdr)
            if (m := re.match(r"^(.+?)\s+(\d{4})$", str(c or "").strip())) and m.group(1).strip() in AMS10]
    section = None
    for r in rows[6:]:
        b = r[0]
        if b is None or str(b).strip() == "": continue
        s = str(b); depth = (len(s)-len(s.lstrip()))//3; name = s.strip()
        if depth == 0:
            section = name if name in ("Primary","Secondary") else None
            continue
        if section is None or depth != 1 or name == "Total": continue
        for j, region, year in cols:
            v = r[j] if j < len(r) else None
            if v in (None,"") or (isinstance(v,str) and not v.strip()): continue
            try: val = float(v)
            except (TypeError,ValueError): continue
            res_rows.append({"variable": variable, "scenario": scenario,
                "energy_class": section, "fuel": name, "region": region,
                "year": year, "value": val, "unit": unit})
RES_FIELDS = ["variable","scenario","energy_class","fuel","region","year","value","unit"]
with open(R/"aeo9_v0.71_resources_supply_exports_tidy.csv","w",newline="",encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=RES_FIELDS); w.writeheader(); w.writerows(res_rows)
print(f"resources tidy: {len(res_rows)} rows")

# ---------- shared full-results set ----------
FULL = [
    (R/"aeo9_v0.71_demand_by_fuel_tidy.csv", "aeo9_v0.71_demand_ALL_sectors_by_fuel.csv"),
    (R/"aeo9_v0.71_supply_power_tidy.csv",   "aeo9_v0.71_supply_power.csv"),
    (R/"aeo9_v0.71_resources_supply_exports_tidy.csv", "aeo9_v0.71_resources_supply_exports.csv"),
    (R/"README_full_results_packaged.md",   "README_full_results.md"),
]

def demand_slice(sectors):
    rows = [r for r in csv.DictReader(open(R/"aeo9_v0.71_demand_by_fuel_tidy.csv",encoding="utf-8"))
            if r["sector"] in sectors]
    return rows, list(rows[0].keys())

# ---------- per-team config ----------
# each: results = ('demand',[sectors]) | ('resources',None) | ('keys',None)
# files map dst-subfolder -> list of (src path relative to inject/, dst name)
TEAMS = {
 "industry": {"results": ("demand", ["Industry"]),
   "sh": "industry/structure_handover_20260703",
   "input": [("industry/structure_handover_20260703/current_expressions_industry_4scenarios.csv","industry_current_expressions.csv")],
   "structure": ["industry_tree.txt","industry_branch_variables_units.csv","README_INDUSTRY_CANON_STRUCTURE.md","ANOMALY_AUDIT_INDUSTRY_20260704.md"],
   "drivers": ["current_expressions_keys_slice_4scenarios.csv","keys_slice_industry.txt","keys_slice_industry_units.csv","current_expressions_resources_slice_4scenarios.csv","resources_slice_industry_units.csv"]},
 "commercial": {"results": ("demand", ["Commercial"]),
   "sh": "commercial/structure_handover_20260703",
   "input": [("commercial/structure_handover_20260703/current_expressions_commercial_4scenarios.csv","commercial_current_expressions.csv")],
   "structure": ["commercial_tree.txt","commercial_branch_variables_units.csv","README_COMMERCIAL_CANON_STRUCTURE.md","ANOMALY_AUDIT_COMMERCIAL_20260704.md"],
   "drivers": ["current_expressions_keys_slice_4scenarios.csv","keys_slice_commercial.txt","keys_slice_commercial_units.csv","current_expressions_resources_slice_4scenarios.csv","resources_slice_commercial_units.csv"]},
 "transport": {"results": ("demand", ["Transport","International Transport"]),
   "sh": "transport/structure_handover_20260703",
   "input": [("transport/canonical_leap_inputs.csv","transport_canonical_input.csv"),
             ("transport/structure_handover_20260703/current_expressions_transport_4scenarios.csv","transport_current_expressions.csv")],
   "structure": ["transport_tree.txt","transport_branch_variables_units.csv","README_TRANSPORT_CANON_STRUCTURE.md","ANOMALY_AUDIT_TRANSPORT_20260704.md"],
   "drivers": ["current_expressions_keys_slice_4scenarios.csv","keys_slice_transport.txt","keys_slice_transport_units.csv","resources_branch_variables_units.csv"]},
 "fossil": {"results": ("resources", None),
   "sh": "fossil/structure_handover_20260703",
   "input": [("fossil/canonical_leap_inputs.csv","fossil_canonical_input.csv"),
             ("fossil/CSV_AUTHORING_GUIDE.md","CSV_AUTHORING_GUIDE.md"),
             ("fossil/structure_handover_20260703/current_expressions_resources_4scenarios.csv","fossil_resources_expressions.csv"),
             ("fossil/structure_handover_20260703/current_expressions_transformation_slice_4scenarios.csv","fossil_transformation_expressions.csv")],
   "structure": ["resources_tree.txt","resources_slice_fossil_units.csv","transformation_slice_tree.txt","transformation_slice_branch_variables_units.csv","README_FOSSIL_CANON_STRUCTURE.md","ANOMALY_AUDIT_FOSSIL_20260704.md"],
   "drivers": []},
 "bioenergy": {"results": ("resources", None),
   "sh": "bioenergy/structure_handover_20260703",
   "input": [("bioenergy/bioenergy_leap_input.csv","bioenergy_input.csv"),
             ("bioenergy/CSV_AUTHORING_GUIDE.md","CSV_AUTHORING_GUIDE.md"),
             ("bioenergy/structure_handover_20260703/current_expressions_resources_4scenarios.csv","bioenergy_resources_expressions.csv"),
             ("bioenergy/structure_handover_20260703/current_expressions_transformation_slice_4scenarios.csv","bioenergy_transformation_expressions.csv")],
   "structure": ["resources_tree.txt","resources_branch_variables_units.csv","transformation_slice_tree.txt","transformation_slice_branch_variables_units.csv","README_BIOENERGY_CANON_STRUCTURE.md","ANOMALY_AUDIT_BIOENERGY_20260704.md"],
   "drivers": ["current_expressions_keys_slice_4scenarios.csv","keys_slice_bioenergy.txt","keys_slice_bioenergy_units.csv"]},
 "keys": {"results": ("keys", None),
   "sh": "keys/structure_handover_20260703",
   "input": [("keys/structure_handover_20260703/current_expressions_keys_4scenarios.csv","keys_current_expressions.csv")],
   "structure": ["keys_tree.txt","keys_branch_variables_units.csv","README_KEYS_CANON_STRUCTURE.md","ANOMALY_AUDIT_KEYS_20260704.md"],
   "drivers": []},
}

def cp(src, dst):
    dst.parent.mkdir(parents=True, exist_ok=True)
    if src.exists(): shutil.copy2(src, dst); return True
    print(f"    MISSING: {src}"); return False

manifest = {}
for team, cfg in TEAMS.items():
    stage = R / f"_stage_{team}"
    if stage.exists(): shutil.rmtree(stage)
    stage.mkdir(parents=True)
    n_res = 0
    # results
    kind, arg = cfg["results"]
    if kind == "demand":
        rows, flds = demand_slice(arg)
        with open(stage/f"{team.upper()}_RESULTS_v0.71.csv","w",newline="",encoding="utf-8") as fh:
            w=csv.DictWriter(fh,fieldnames=flds); w.writeheader(); w.writerows(rows); n_res=len(rows)
    elif kind == "resources":
        BIO={"Cassava","Coconut Oil","Palm Oil","Sugarcane","Molasses","Biomass","Bagasse","Wood","Biogas","Domestic Biogas","Biodiesel","Ethanol","Biomethane","Sustainable Aviation Fuel","SAF","Charcoal","Other Biomass","Efficient Wood","Municipal Solid Waste"}
        FOSSILF={"Coal Anthracite","Coal Bituminous","Coal Lignite","Coal Sub bituminous","Coal Unspecified","Coal","Crude Oil","Natural Gas","LNG","Diesel","Gasoline","Aviation Gasoline","Avgas","Blended Gasoline","Blended Diesel","Kerosene","Jet Kerosene","Residual Fuel Oil","LPG","Bitumen","Petroleum Coke","Refinery Gas","Naphtha","Lubricants","Oil","Metalurgical Coke","Hard Coal Briquettes","Brown Coal Briquettes","Coke Oven Gas","Blast Furnace Gas"}
        keep = BIO if team=="bioenergy" else FOSSILF
        sub=[r for r in res_rows if r["fuel"] in keep]
        with open(stage/f"{team.upper()}_RESULTS_v0.71_resources.csv","w",newline="",encoding="utf-8") as fh:
            w=csv.DictWriter(fh,fieldnames=RES_FIELDS); w.writeheader(); w.writerows(sub)
        n_res=len(sub)
    else:  # keys: no direct energy result; deliverable is the driver expressions
        cp(H/cfg["sh"]/"current_expressions_keys_4scenarios.csv", stage/"KEYS_driver_values_v0.71.csv")
    # input
    for src, name in cfg["input"]:
        cp(H/src, stage/"input"/name)
    # leap_structure
    for f in cfg["structure"]:
        cp(H/cfg["sh"]/f, stage/"leap_structure"/f)
    # connected_drivers
    for f in cfg["drivers"]:
        cp(H/cfg["sh"]/f, stage/"connected_drivers"/f)
    # full_results
    for src,name in FULL:
        cp(src, stage/"full_results"/name)
    # zip
    dst = OUTBOX/f"{team}_v071_results_20260709.zip"
    with zipfile.ZipFile(dst,"w",zipfile.ZIP_DEFLATED) as z:
        for root,_,files in os.walk(stage):
            for fn in files:
                full=Path(root)/fn; z.write(full, full.relative_to(stage))
    names=[n for n in zipfile.ZipFile(dst).namelist()]
    manifest[team]={"result_rows":n_res,"files":len(names),"kb":round(dst.stat().st_size/1024,1),"names":names}

print("\n=== PACKAGE MANIFEST ===")
for t,m in manifest.items():
    print(f"\n{t}: {m['files']} files, {m['kb']} KB, result_rows={m['result_rows']}")
    for n in sorted(m["names"]):
        if not n.startswith("full_results"): print(f"    {n}")
